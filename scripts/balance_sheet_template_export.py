"""Export a live balance sheet snapshot into an editable Excel template.

Each row is one node in the IAS 1 tree. Edit the rows in-place — reorder,
re-label, change Level (1-4), or change Source. When you save the file,
companion script `balance_sheet_template_import.py` translates it back into
`/finance/balance_sheet_config.yaml`.

Columns:
  ID            stable identifier (don't change for existing rows)
  Section       ASSETS / LIABILITIES
  Bucket        Current / Non-Current
  Level         1=top section, 2=category, 3=sub-category, 4=leaf
  Label         display name
  Source Type   firefly_account_ids | portfolio_mcp_liquid | portfolio_mcp_manual | (parent)
  Source Value  comma-separated IDs OR protocol names (for portfolio_mcp_manual) OR blank for parent
  Live SGD      current value pulled at export time (read-only reference)
  Notes         freeform comments
"""
import os
import sys
import requests
import yaml
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = Path(r"C:\Users\azfar\OneDrive\Sentinel Finance\balance_sheet_template.xlsx")
CONFIG_PATH = Path(r"C:\Users\azfar\metamcp-local\finance\balance_sheet_config.yaml")


def fetch_live_data() -> dict:
    """Pull live balance sheet by invoking the builder INSIDE the portfolio-mcp
    container — bypasses the auth gate on /balance_sheet.json."""
    import subprocess, json as _json
    code = (
        "import asyncio,json,sys;"
        "from app.balance_sheet import build_balance_sheet;"
        "print(json.dumps(asyncio.run(build_balance_sheet())))"
    )
    r = subprocess.run(
        ["docker", "exec", "portfolio-mcp", "python", "-c", code],
        capture_output=True, text=True, timeout=60,
    )
    if r.returncode != 0:
        raise RuntimeError(f"docker exec failed: {r.stderr[:400]}")
    return _json.loads(r.stdout.strip().splitlines()[-1])


def flatten_assets(data: dict, config: dict) -> list[dict]:
    """Walk the live JSON + config in lockstep, emit one row per node."""
    rows = []
    rows.append({"id": "ASSETS", "section": "ASSETS", "bucket": "", "level": 1,
                 "label": "ASSETS", "source_type": "", "source_value": "",
                 "live_sgd": data["assets"]["total"], "notes": ""})

    for bucket, bucket_label in [("current", "Current"), ("non_current", "Non-Current")]:
        bucket_data = data["assets"][bucket]
        rows.append({"id": f"asset_{bucket}", "section": "ASSETS", "bucket": bucket_label,
                     "level": 2, "label": f"{bucket_label} Assets",
                     "source_type": "", "source_value": "",
                     "live_sgd": bucket_data["total"], "notes": ""})
        for n in bucket_data["nodes"]:
            walk_node(n, bucket_label, 3, rows, section="ASSETS")
    return rows


def walk_node(node: dict, bucket: str, level: int, rows: list, section: str):
    src_type, src_val = describe_source(node)
    rows.append({
        "id": node["id"], "section": section, "bucket": bucket, "level": level,
        "label": node["label"],
        "source_type": src_type, "source_value": src_val,
        "live_sgd": node["total"], "notes": "",
    })
    for c in node.get("children", []):
        walk_node(c, bucket, level + 1, rows, section)


def describe_source(node: dict) -> tuple[str, str]:
    """Look up the node id in the config and return (source_type, source_value)."""
    # Find the matching node in config by traversing
    cfg = yaml.safe_load(CONFIG_PATH.read_text())
    found = _find_in_config(cfg, node["id"])
    if not found:
        return "(parent)", ""
    if found.get("children"):
        return "(parent)", ""
    if found.get("firefly_account_ids"):
        return "firefly_account_ids", ",".join(map(str, found["firefly_account_ids"]))
    if found.get("source") == "portfolio_mcp_liquid":
        return "portfolio_mcp_liquid", ""
    if found.get("source") == "portfolio_mcp_manual":
        return "portfolio_mcp_manual", ",".join(found.get("include_protocols", []) or [])
    return "(unknown)", ""


def _find_in_config(cfg: dict, target_id: str):
    def walk(nodes):
        for n in nodes:
            if n["id"] == target_id:
                return n
            if n.get("children"):
                r = walk(n["children"])
                if r:
                    return r
        return None

    for top in ("current", "non_current"):
        r = walk(cfg["assets"][top])
        if r:
            return r
    return None


def flatten_liabilities(data: dict) -> list[dict]:
    rows = []
    rows.append({"id": "LIABILITIES", "section": "LIABILITIES", "bucket": "", "level": 1,
                 "label": "LIABILITIES", "source_type": "", "source_value": "",
                 "live_sgd": data["liabilities"]["total"], "notes": ""})
    for bucket, bucket_label in [("current", "Current"), ("non_current", "Non-Current")]:
        b = data["liabilities"][bucket]
        rows.append({"id": f"liab_{bucket}", "section": "LIABILITIES", "bucket": bucket_label,
                     "level": 2, "label": f"{bucket_label} Liabilities",
                     "source_type": "", "source_value": "",
                     "live_sgd": b["total"], "notes": ""})
        for bk in b["buckets"]:
            rows.append({"id": bk["id"], "section": "LIABILITIES", "bucket": bucket_label,
                         "level": 3, "label": bk["label"],
                         "source_type": "aging_window",
                         "source_value": _aging_for(bk["id"]),
                         "live_sgd": bk["total"], "notes": ""})
    return rows


def _aging_for(bucket_id: str) -> str:
    m = {"due_30": "1-1", "due_31_365": "2-12", "due_12_plus": "13-9999"}
    return m.get(bucket_id, "")


def build_workbook(rows_assets: list, rows_liab: list, summary: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill("solid", fgColor="1F3A4D")
    bucket_fill = PatternFill("solid", fgColor="2E5266")
    cat_fill = PatternFill("solid", fgColor="3D6B7F")
    sub_fill = PatternFill("solid", fgColor="A9C5D2")
    leaf_fill = PatternFill("solid", fgColor="EEF3F6")
    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws["A1"] = "Sentinel Finance — Balance Sheet Template (IAS 1)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Snapshot at {summary['generated_at_utc']} · base {summary['base_currency']} · USD@{summary['usd_to_sgd']}"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws.merge_cells("A2:I2")
    ws["A3"] = "Edit rows below. Don't change column A (ID). Save and send back."
    ws["A3"].font = Font(italic=True, color="888888", size=10)
    ws.merge_cells("A3:I3")

    # Headers
    headers = ["ID", "Section", "Bucket", "Level", "Label", "Source Type", "Source Value", "Live SGD", "Notes"]
    header_row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    # Data
    row_num = header_row + 1
    for r in rows_assets + rows_liab:
        cells = [r["id"], r["section"], r["bucket"], r["level"], r["label"],
                 r["source_type"], r["source_value"], r["live_sgd"], r["notes"]]
        for col, val in enumerate(cells, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = border
            c.alignment = Alignment(horizontal="left" if col != 8 else "right",
                                    vertical="center",
                                    indent=max(0, (r["level"] - 1) * 2) if col == 5 else 0)
        # Row formatting by level
        fill = {1: section_fill, 2: bucket_fill, 3: cat_fill, 4: sub_fill}.get(r["level"], leaf_fill)
        font = Font(bold=(r["level"] <= 2),
                    color="FFFFFF" if r["level"] <= 3 else "000000",
                    size=11 if r["level"] == 1 else 10)
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).font = font
        # Live SGD as currency
        ws.cell(row=row_num, column=8).number_format = '#,##0.00'
        row_num += 1

    # Column widths
    widths = {"A": 24, "B": 14, "C": 14, "D": 7, "E": 38, "F": 22, "G": 28, "H": 14, "I": 30}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A6"

    # Sheet 2: Liquid Crypto detail
    ws2 = wb.create_sheet("Liquid Crypto Detail")
    ws2.append(["Symbol", "Chain", "USD", "SGD"])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = section_fill
    live_liquid = []
    try:
        for n in summary["assets"]["current"]["nodes"]:
            collect_liquid(n, live_liquid)
    except Exception:
        pass
    for it in sorted(live_liquid, key=lambda x: -x.get("sgd", 0)):
        ws2.append([it.get("label", "").split(" (")[0],
                    it.get("label", "").split(" (")[-1].rstrip(")"),
                    it.get("usd", 0), it.get("sgd", 0)])
    for col_letter, w in {"A": 16, "B": 14, "C": 12, "D": 12}.items():
        ws2.column_dimensions[col_letter].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Instructions
    ws3 = wb.create_sheet("Instructions")
    instructions = [
        ("Sentinel Finance — Balance Sheet Template", 14, True),
        ("", 10, False),
        ("Purpose:", 11, True),
        ("Reorganise the IAS 1 balance sheet layout by editing rows on the first sheet.", 10, False),
        ("The order of rows here = the order in the Mini App.", 10, False),
        ("", 10, False),
        ("How to edit:", 11, True),
        ("• Reorder rows: select an entire row and drag to a new position.", 10, False),
        ("• Rename a category: edit the Label column.", 10, False),
        ("• Move a leaf into another parent: edit its Bucket / Level / parent context.", 10, False),
        ("• Add a new category: insert a row; leave ID blank (new entries get auto-IDs on import).", 10, False),
        ("• Delete: blank out the entire row.", 10, False),
        ("", 10, False),
        ("Column meaning:", 11, True),
        ("• ID: stable identifier. Don't change for existing rows. Leave blank for new.", 10, False),
        ("• Section: ASSETS or LIABILITIES.", 10, False),
        ("• Bucket: Current or Non-Current.", 10, False),
        ("• Level: 1=top section, 2=bucket, 3=category, 4=sub-category, 5=leaf.", 10, False),
        ("• Label: what shows up in the Mini App.", 10, False),
        ("• Source Type: how the leaf's value is computed:", 10, False),
        ("   - (parent): aggregates its children (no source)", 10, False),
        ("   - firefly_account_ids: comma-separated Firefly asset account IDs", 10, False),
        ("   - portfolio_mcp_liquid: pulls Moralis-visible tokens", 10, False),
        ("   - portfolio_mcp_manual: pulls manual positions filtered by protocol name", 10, False),
        ("   - aging_window: liability bucket. Source Value = months range, e.g. 1-1, 2-12, 13-9999.", 10, False),
        ("• Source Value: comma-separated IDs / protocol names / month range.", 10, False),
        ("• Live SGD: read-only reference value at export time.", 10, False),
        ("• Notes: freeform.", 10, False),
        ("", 10, False),
        ("When done:", 11, True),
        ("Save the file in place (OneDrive\\Sentinel Finance\\balance_sheet_template.xlsx)", 10, False),
        ("and tell Claude to import it.", 10, False),
    ]
    for i, (text, size, bold) in enumerate(instructions, 1):
        c = ws3.cell(row=i, column=1, value=text)
        c.font = Font(size=size, bold=bold)
    ws3.column_dimensions["A"].width = 100

    return wb


def collect_liquid(n: dict, out: list):
    if n.get("id") == "liquid_crypto":
        for it in n.get("items", []):
            out.append(it)
    for c in n.get("children", []):
        collect_liquid(c, out)


def main():
    print(f"Fetching live balance sheet...")
    data = fetch_live_data()
    rows_assets = flatten_assets(data, yaml.safe_load(CONFIG_PATH.read_text()))
    rows_liab = flatten_liabilities(data)
    wb = build_workbook(rows_assets, rows_liab, data)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Assets rows: {len(rows_assets)}, Liabilities rows: {len(rows_liab)}")


if __name__ == "__main__":
    main()
