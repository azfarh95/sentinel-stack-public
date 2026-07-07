"""Balance Sheet template v2 — matches user's hand-drawn structure.

ASSETS / Current
  Cash & Bank          (POSB only)
  Crypto Wallets       (Coinbase, Crypto.com)
  DeFi                 (Lending Protocols, LP Positions, Yield aggregators / vaults)
  Token Holdings       (BNB, Base, Cronos named; >$50 others bucket; dust bucket)

ASSETS / Non-Current
  CPF                  (IS, MA, OA, SA)
  ILP                  (Singlife Savvy Invest, Tokio Marine)  — TODO
  Other Cash Investments (iFast, Tiger Brokers)              — TODO
  Staking Vaults       (Wolfswap)

LIABILITIES
  Current   — Due within 30 days, Due within 31 to 365 days
  Non-Current — Due > 12 months
  Per-bucket account breakdown alphabetised.

Columns: ID | Section | Bucket | Level | Label | Source Type | Source Value | USD | SGD | Notes
"""
import json
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_PATH = Path(r"C:\Users\azfar\OneDrive\Sentinel Finance\balance_sheet_template_v2.xlsx")
CHAIN_THRESHOLD_USD = 50.0
NAMED_CHAINS = {"bsc": "BNB Chain", "base": "Base Chain", "cronos": "Cronos Chain"}


def fetch_live_data() -> dict:
    code = (
        "import asyncio,json,sys;"
        "from app.balance_sheet import build_balance_sheet;"
        "print(json.dumps(asyncio.run(build_balance_sheet())))"
    )
    r = subprocess.run(["docker", "exec", "portfolio-mcp", "python", "-c", code],
                       capture_output=True, text=True, timeout=60)
    if r.returncode != 0:
        raise RuntimeError(f"docker exec failed: {r.stderr[:400]}")
    return json.loads(r.stdout.strip().splitlines()[-1])


def fetch_liquid_positions() -> list[dict]:
    """Pull raw positions list (one entry per token) inside the container."""
    code = (
        "import asyncio,json;"
        "from app.main import portfolio_snapshot;"
        "snap = asyncio.run(portfolio_snapshot(None, save=False));"
        "print(json.dumps(snap.get('positions', [])))"
    )
    r = subprocess.run(["docker", "exec", "portfolio-mcp", "python", "-c", code],
                       capture_output=True, text=True, timeout=120)
    if r.returncode != 0:
        raise RuntimeError(f"docker exec snapshot failed: {r.stderr[:400]}")
    return json.loads(r.stdout.strip().splitlines()[-1])


def fetch_liab_breakdown(data: dict) -> dict:
    """Returns {bucket_id: [{name, sgd}, ...] sorted alphabetically}."""
    out = {}
    for b in data["liabilities"]["current"]["buckets"] + data["liabilities"]["non_current"]["buckets"]:
        items = sorted(b["breakdown"], key=lambda x: x["name"].lower())
        out[b["id"]] = items
    return out


def group_tokens_by_chain(positions: list[dict]) -> dict:
    """Returns {chain: total_usd, ...} for all chains."""
    by_chain = defaultdict(float)
    for p in positions:
        by_chain[p["chain"]] += p["usd_value"]
    return dict(by_chain)


def build_rows(data: dict) -> tuple[list, dict, list]:
    """Returns (asset_rows, summary, liab_rows)."""
    fx = float(data.get("usd_to_sgd", 1.34))

    positions = fetch_liquid_positions()
    chains_total = group_tokens_by_chain(positions)

    # Categorise chains
    named_usd = {key: chains_total.get(key, 0.0) for key in NAMED_CHAINS}
    other_above = {k: v for k, v in chains_total.items()
                   if k not in NAMED_CHAINS and v >= CHAIN_THRESHOLD_USD}
    dust = {k: v for k, v in chains_total.items()
            if k not in NAMED_CHAINS and v < CHAIN_THRESHOLD_USD}

    # Pre-compute live values
    cash_posb_sgd = _firefly_value(data, [1])
    cex_coinbase_sgd = _firefly_value(data, [97])
    cex_cryptocom_sgd = _firefly_value(data, [98])
    cpf_is_sgd = _firefly_value(data, [147])
    cpf_oa_sgd = _firefly_value(data, [141])
    cpf_sa_sgd = _firefly_value(data, [143])
    cpf_ma_sgd = _firefly_value(data, [145])

    # Wolfswap (only manual position currently)
    wolfswap_node = _find_node(data["assets"]["non_current"]["nodes"], "staking_vaults")
    wolfswap_usd = (wolfswap_node or {}).get("total", 0.0) / fx if wolfswap_node else 0.0
    wolfswap_sgd = (wolfswap_node or {}).get("total", 0.0) if wolfswap_node else 0.0

    rows = []
    next_id = [0]
    def aid(prefix): next_id[0] += 1; return f"{prefix}_{next_id[0]:02d}"

    def add(section, bucket, level, label, src_type, src_val, usd, sgd, notes=""):
        rows.append({
            "id": aid(section[:3].lower() + "_" + (bucket[:3].lower() if bucket else "all")),
            "section": section, "bucket": bucket, "level": level, "label": label,
            "source_type": src_type, "source_value": src_val,
            "usd": round(usd, 2) if usd else 0.0, "sgd": round(sgd, 2) if sgd else 0.0,
            "notes": notes,
        })

    # ── ASSETS ──
    add("ASSETS", "", 1, "ASSETS", "", "", 0, 0)
    add("ASSETS", "Current", 2, "Current Assets", "(parent)", "", 0, 0)

    # Cash & Bank
    add("ASSETS", "Current", 3, "Cash & Bank", "(parent)", "", 0, 0)
    add("ASSETS", "Current", 4, "POSB Savings", "firefly_account_ids", "1",
        cash_posb_sgd / fx, cash_posb_sgd)

    # Crypto Wallets
    add("ASSETS", "Current", 3, "Crypto Wallets", "(parent)", "", 0, 0)
    add("ASSETS", "Current", 4, "Coinbase Account", "firefly_account_ids", "97",
        cex_coinbase_sgd / fx, cex_coinbase_sgd)
    add("ASSETS", "Current", 4, "Crypto.com Account", "firefly_account_ids", "98",
        cex_cryptocom_sgd / fx, cex_cryptocom_sgd)

    # DeFi
    add("ASSETS", "Current", 3, "DeFi", "(parent)", "", 0, 0)
    add("ASSETS", "Current", 4, "Lending Protocols", "portfolio_mcp_manual", "(none yet)",
        0, 0, "TODO: tag manual positions with protocol type 'lending'")
    add("ASSETS", "Current", 4, "LP Positions", "portfolio_mcp_manual", "(none yet)",
        0, 0, "TODO: tag manual positions with protocol type 'lp'")
    add("ASSETS", "Current", 4, "Yield Aggregators / Vaults", "portfolio_mcp_manual", "(none yet)",
        0, 0, "TODO: e.g. Beefy, Yearn, autocompounders")

    # Token Holdings — alphabetically: Base, BNB, Cronos
    add("ASSETS", "Current", 3, "Token Holdings", "(parent)", "", 0, 0)
    for chain_slug in sorted(NAMED_CHAINS.keys(), key=lambda c: NAMED_CHAINS[c]):
        chain_label = NAMED_CHAINS[chain_slug]
        chain_usd = named_usd[chain_slug]
        add("ASSETS", "Current", 4, chain_label, "portfolio_mcp_liquid_chain", chain_slug,
            chain_usd, chain_usd * fx)
    other_above_usd = sum(other_above.values())
    add("ASSETS", "Current", 4, f"Other Chains (>${int(CHAIN_THRESHOLD_USD)})",
        "portfolio_mcp_liquid_other_above", ",".join(sorted(other_above.keys())) or "(none)",
        other_above_usd, other_above_usd * fx)
    dust_usd = sum(dust.values())
    add("ASSETS", "Current", 4, f"Dust Chains (<${int(CHAIN_THRESHOLD_USD)})",
        "portfolio_mcp_liquid_dust", ",".join(sorted(dust.keys())) or "(none)",
        dust_usd, dust_usd * fx)

    # ── NON-CURRENT ──
    add("ASSETS", "Non-Current", 2, "Non-Current Assets", "(parent)", "", 0, 0)

    # CPF
    add("ASSETS", "Non-Current", 3, "CPF", "(parent)", "", 0, 0)
    for label, ff_id, sgd in sorted([
        ("CPF IS", 147, cpf_is_sgd),
        ("CPF MA", 145, cpf_ma_sgd),
        ("CPF OA", 141, cpf_oa_sgd),
        ("CPF SA", 143, cpf_sa_sgd),
    ], key=lambda x: x[0]):
        add("ASSETS", "Non-Current", 4, label, "firefly_account_ids", str(ff_id),
            sgd / fx, sgd)

    # ILP — TODO placeholders
    add("ASSETS", "Non-Current", 3, "ILP", "(parent)", "", 0, 0, "TODO: insurance-linked plans")
    add("ASSETS", "Non-Current", 4, "Singlife Savvy Invest", "(todo)", "", 0, 0,
        "TODO: add Firefly asset account + manual valuation")
    add("ASSETS", "Non-Current", 4, "Tokio Marine", "(todo)", "", 0, 0,
        "TODO: add Firefly asset account + manual valuation")

    # Other Cash Investments — TODO
    add("ASSETS", "Non-Current", 3, "Other Cash Investments", "(parent)", "", 0, 0,
        "TODO: brokerage holdings not on iFAST")
    add("ASSETS", "Non-Current", 4, "iFast", "(todo)", "", 0, 0,
        "TODO: SRS / cash account at iFAST (non-CPF)")
    add("ASSETS", "Non-Current", 4, "Tiger Brokers", "(todo)", "", 0, 0,
        "TODO: brokerage holdings")

    # Staking Vaults
    add("ASSETS", "Non-Current", 3, "Staking Vaults", "(parent)", "", 0, 0)
    add("ASSETS", "Non-Current", 4, "Wolfswap", "portfolio_mcp_manual", "WolfSwap",
        wolfswap_usd, wolfswap_sgd)

    # Asset totals row
    total_assets_usd = sum(r["usd"] for r in rows if r["level"] == 4 and r["section"] == "ASSETS")
    total_assets_sgd = sum(r["sgd"] for r in rows if r["level"] == 4 and r["section"] == "ASSETS")
    # Fix level-1 row to show totals
    rows[0]["usd"] = round(total_assets_usd, 2)
    rows[0]["sgd"] = round(total_assets_sgd, 2)

    # ── LIABILITIES ──
    liab_break = fetch_liab_breakdown(data)
    add("LIABILITIES", "", 1, "LIABILITIES", "", "",
        data["liabilities"]["total"] / fx, data["liabilities"]["total"])
    add("LIABILITIES", "Current", 2, "Current Liabilities", "(parent)", "", 0, 0)
    for b in data["liabilities"]["current"]["buckets"]:
        add("LIABILITIES", "Current", 3, b["label"], "aging_window", _aging(b["id"]),
            b["total"] / fx, b["total"])
        for it in liab_break.get(b["id"], []):
            add("LIABILITIES", "Current", 4, it["name"], "registry_account", "",
                it["sgd"] / fx, it["sgd"])
    add("LIABILITIES", "Non-Current", 2, "Non-Current Liabilities", "(parent)", "", 0, 0)
    for b in data["liabilities"]["non_current"]["buckets"]:
        add("LIABILITIES", "Non-Current", 3, b["label"], "aging_window", _aging(b["id"]),
            b["total"] / fx, b["total"])
        for it in liab_break.get(b["id"], []):
            add("LIABILITIES", "Non-Current", 4, it["name"], "registry_account", "",
                it["sgd"] / fx, it["sgd"])

    return rows, data, []


def _firefly_value(data: dict, account_ids: list[int]) -> float:
    """Walk the tree for items matching given Firefly account IDs."""
    matched = []
    def walk(nodes):
        for n in nodes:
            for it in n.get("items", []) or []:
                # Items don't carry the FF id directly; match by SGD-balance heuristic
                # by walking the tree's structure (firefly_account_ids set in config).
                # Simpler: pass through to a known map below.
                pass
            walk(n.get("children", []))
    # Static map from the original config (asset id -> live SGD)
    label_map = {}
    def collect(nodes):
        for n in nodes:
            for it in n.get("items", []) or []:
                if "label" in it and "currency" in it and "sgd" in it:
                    label_map[it["label"]] = it["sgd"]
            collect(n.get("children", []))
    collect(data["assets"]["current"]["nodes"])
    collect(data["assets"]["non_current"]["nodes"])
    # Map our known IDs to labels
    id_to_label = {
        1: "POSB Savings", 4: "Cash wallet", 97: "Coinbase Account",
        98: "Crypto.com Account", 141: "CPF OA", 143: "CPF SA",
        145: "CPF MA", 147: "CPF Investment Scheme",
    }
    total = 0.0
    for aid in account_ids:
        lbl = id_to_label.get(aid)
        if lbl and lbl in label_map:
            total += label_map[lbl]
    return total


def _find_node(nodes: list, target_id: str):
    for n in nodes:
        if n.get("id") == target_id:
            return n
        c = _find_node(n.get("children", []), target_id)
        if c:
            return c
    return None


def _aging(bucket_id: str) -> str:
    return {"due_30": "1-1", "due_31_365": "2-12", "due_12_plus": "13-9999"}.get(bucket_id, "")


def build_workbook(rows: list, summary: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    section_fill = PatternFill("solid", fgColor="1F3A4D")
    bucket_fill = PatternFill("solid", fgColor="2E5266")
    cat_fill = PatternFill("solid", fgColor="3D6B7F")
    leaf_fill = PatternFill("solid", fgColor="EEF3F6")
    todo_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fx = summary.get("usd_to_sgd", 1.34)
    ws["A1"] = "Sentinel Finance — Balance Sheet (IAS 1)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A2"] = f"Snapshot {summary['generated_at_utc']} · base SGD · USD@{fx} (xe.com — edit in config)"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws.merge_cells("A2:J2")
    ws["A3"] = "TODO rows = data sources not wired up yet. Yellow = action needed."
    ws["A3"].font = Font(italic=True, color="888888", size=10)
    ws.merge_cells("A3:J3")

    headers = ["ID", "Section", "Bucket", "Level", "Label", "Source Type", "Source Value", "USD", "SGD", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    row_num = 6
    for r in rows:
        vals = [r["id"], r["section"], r["bucket"], r["level"], r["label"],
                r["source_type"], r["source_value"], r["usd"], r["sgd"], r["notes"]]
        is_todo = "TODO" in (r["notes"] or "") or r["source_type"] in ("(todo)",)
        fill = todo_fill if is_todo else (
            {1: section_fill, 2: bucket_fill, 3: cat_fill}.get(r["level"], leaf_fill)
        )
        font_color = "000000" if r["level"] >= 4 or is_todo else "FFFFFF"
        font_bold = r["level"] <= 3
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = border
            c.fill = fill
            c.font = Font(color=font_color, bold=font_bold,
                          size=11 if r["level"] == 1 else 10)
            if col == 5:
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=max(0, r["level"] - 1) * 2)
            elif col in (8, 9):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

    widths = {"A": 14, "B": 12, "C": 12, "D": 6, "E": 36, "F": 26, "G": 24, "H": 12, "I": 12, "J": 30}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A6"

    # Sheet 2: Liquid Crypto Detail (sorted by USD desc)
    ws2 = wb.create_sheet("Liquid Crypto Detail")
    ws2.append(["Symbol", "Chain", "USD", "SGD"])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = section_fill
    try:
        positions = fetch_liquid_positions()
        for p in sorted(positions, key=lambda x: -x["usd_value"]):
            ws2.append([p["symbol"], p["chain"], round(p["usd_value"], 2),
                        round(p["usd_value"] * fx, 2)])
    except Exception as e:
        ws2.append([f"(error fetching positions: {e})", "", 0, 0])
    for letter, w in {"A": 16, "B": 14, "C": 12, "D": 12}.items():
        ws2.column_dimensions[letter].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Config notes
    ws3 = wb.create_sheet("Config")
    config_rows = [
        ("Setting", "Current Value", "Notes"),
        ("USD → SGD FX rate", fx, "Hardcoded in finance/balance_sheet_config.yaml. xe.com source. Future: settings page."),
        ("Chain threshold (USD)", CHAIN_THRESHOLD_USD, "Chains with total > this = named row. Lower = dust."),
        ("Named chains", ", ".join(NAMED_CHAINS.values()), "Always shown separately regardless of value."),
    ]
    for r_idx, row in enumerate(config_rows, 1):
        for c_idx, val in enumerate(row, 1):
            c = ws3.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = section_fill
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 80

    return wb


def main():
    print("Fetching live balance sheet + positions...")
    data = fetch_live_data()
    rows, summary, _ = build_rows(data)
    wb = build_workbook(rows, summary)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Total rows: {len(rows)}")
    todos = sum(1 for r in rows if "TODO" in (r["notes"] or "") or r["source_type"] == "(todo)")
    print(f"  TODO rows: {todos}")


if __name__ == "__main__":
    main()
